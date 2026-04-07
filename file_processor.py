"""
Advanced File Processing for Threat Modeling
Supports PDF, DOCX, XLSX, images (PNG, JPG) with OCR
"""

import io
import base64
from typing import Optional, Dict, Any
from pathlib import Path
import logging

logger = logging.getLogger(__name__)


def resize_image_for_api(file_content: bytes, max_dimension: int = 1568) -> bytes:
    """
    Resize image to reduce token consumption while preserving readability.
    Claude's vision processes images at ~1568px max internally, so anything
    larger just wastes tokens without improving analysis quality.
    """
    try:
        from PIL import Image
        img = Image.open(io.BytesIO(file_content))
        original_size = img.size
        
        # Only resize if larger than max_dimension
        if max(img.size) > max_dimension:
            img.thumbnail((max_dimension, max_dimension), Image.LANCZOS)
            buffer = io.BytesIO()
            # Use original format if available, otherwise PNG
            img_format = img.format or 'PNG'
            if img_format.upper() == 'JPEG':
                img.save(buffer, format='JPEG', quality=85)
            else:
                img.save(buffer, format=img_format)
            resized_bytes = buffer.getvalue()
            logger.info(f"🔧 Resized image from {original_size} to {img.size} ({len(file_content)//1024}KB → {len(resized_bytes)//1024}KB)")
            return resized_bytes
        
        logger.info(f"✅ Image {original_size} already within {max_dimension}px limit, no resize needed")
        return file_content
    except ImportError:
        logger.warning("⚠️ Pillow not installed, skipping image resize")
        return file_content
    except Exception as e:
        logger.warning(f"⚠️ Could not resize image: {e}, using original")
        return file_content


def extract_text_from_pdf(file_content: bytes) -> str:
    """Extract text from PDF file"""
    try:
        import PyPDF2
        pdf_file = io.BytesIO(file_content)
        reader = PyPDF2.PdfReader(pdf_file)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"PDF extraction error: {e}")
        return f"[Error extracting PDF: {str(e)}]"


def extract_text_from_docx(file_content: bytes) -> str:
    """Extract text from DOCX file"""
    try:
        from docx import Document
        doc_file = io.BytesIO(file_content)
        doc = Document(doc_file)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text.strip()
    except Exception as e:
        logger.error(f"DOCX extraction error: {e}")
        return f"[Error extracting DOCX: {str(e)}]"


def extract_text_from_xlsx(file_content: bytes) -> str:
    """Extract text from XLSX file"""
    try:
        import openpyxl
        xlsx_file = io.BytesIO(file_content)
        workbook = openpyxl.load_workbook(xlsx_file)
        text = ""
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text += f"\n## Sheet: {sheet_name}\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text.strip()
    except Exception as e:
        logger.error(f"XLSX extraction error: {e}")
        return f"[Error extracting XLSX: {str(e)}]"


def extract_text_from_image_ocr(file_content: bytes) -> str:
    """Extract text from image using OCR (Tesseract)"""
    try:
        from PIL import Image
        import pytesseract
        image = Image.open(io.BytesIO(file_content))
        text = pytesseract.image_to_string(image)
        return text.strip()
    except Exception as e:
        logger.error(f"OCR extraction error: {e}")
        return f"[Error extracting text from image: {str(e)}. Note: Tesseract OCR may not be installed.]"


def encode_image_for_claude(file_content: bytes, file_type: str) -> Dict[str, Any]:
    """Encode image for Claude Vision API"""
    try:
        # Determine media type
        media_type_map = {
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
            'gif': 'image/gif',
            'webp': 'image/webp'
        }
        media_type = media_type_map.get(file_type.lower(), 'image/jpeg')
        
        # Encode to base64
        base64_image = base64.standard_b64encode(file_content).decode('utf-8')
        
        return {
            'type': 'image',
            'source': {
                'type': 'base64',
                'media_type': media_type,
                'data': base64_image
            }
        }
    except Exception as e:
        logger.error(f"Image encoding error: {e}")
        return None


def process_file(filename: str, file_content: bytes, use_vision_api: bool = False, max_chars_per_file: int = 100000) -> str:
    """
    Process any file type and extract text content
    
    Args:
        filename: Name of the file
        file_content: Binary content of the file
        use_vision_api: If True, return image data for Claude Vision API instead of OCR
        max_chars_per_file: Maximum characters to extract per file (default 100k = ~25k tokens)
        
    Returns:
        Extracted text content or placeholder
    """
    try:
        file_extension = Path(filename).suffix.lower().lstrip('.')
        
        # Text files - read directly
        if file_extension in ['txt', 'md', 'csv', 'json', 'xml', 'log']:
            content = file_content.decode('utf-8', errors='ignore')
            return truncate_content(content, max_chars_per_file)
        
        # PDF files
        elif file_extension == 'pdf':
            content = extract_text_from_pdf(file_content)
            return truncate_content(content, max_chars_per_file)
        
        # Word documents
        elif file_extension in ['docx', 'doc']:
            if file_extension == 'docx':
                content = extract_text_from_docx(file_content)
                return truncate_content(content, max_chars_per_file)
            else:
                return f"[.DOC format not supported. Please convert to .DOCX]"
        
        # Excel files
        elif file_extension in ['xlsx', 'xls']:
            if file_extension == 'xlsx':
                content = extract_text_from_xlsx(file_content)
                return truncate_content(content, max_chars_per_file)
            else:
                return f"[.XLS format not supported. Please convert to .XLSX]"
        
        # Image files
        elif file_extension in ['png', 'jpg', 'jpeg', 'gif', 'webp', 'bmp']:
            if use_vision_api:
                # Return special marker for vision API processing
                return f"[IMAGE_FOR_VISION_API: {filename}]"
            else:
                # Use OCR
                ocr_text = extract_text_from_image_ocr(file_content)
                content = f"### Image: {filename}\n[OCR Extracted Text]\n{ocr_text}"
                return truncate_content(content, max_chars_per_file)
        
        # Unsupported formats
        else:
            return f"[{file_extension.upper()} Document: {filename}]"
            
    except Exception as e:
        logger.error(f"File processing error for {filename}: {e}")
        return f"[Error processing {filename}: {str(e)}]"


def smart_truncate(content: str, max_chars: int, filename: str = "") -> str:
    """
    Intelligently truncate content by keeping important sections
    - Keeps beginning (context/overview)
    - Samples middle sections
    - Keeps end (conclusions)
    """
    if len(content) <= max_chars:
        return content
    
    # For very small limits, just take from start
    if max_chars < 1000:
        return content[:max_chars] + f"\n\n[Truncated - showing {max_chars:,} of {len(content):,} chars]"
    
    # Intelligent distribution:
    # 50% from beginning (usually has overview/TOC)
    # 25% from middle (sample content)
    # 25% from end (conclusions/summary)
    beginning_size = int(max_chars * 0.5)
    middle_size = int(max_chars * 0.25)
    end_size = max_chars - beginning_size - middle_size
    
    beginning = content[:beginning_size]
    middle_start = len(content) // 2 - middle_size // 2
    middle = content[middle_start:middle_start + middle_size]
    end = content[-end_size:]
    
    truncation_notice = f"\n\n{'='*50}\n[INTELLIGENT TRUNCATION APPLIED]\nOriginal size: {len(content):,} characters\nShowing: {max_chars:,} characters\nDistribution: 50% beginning, 25% middle sample, 25% end\n{'='*50}\n\n"
    
    return beginning + truncation_notice + "... [MIDDLE SECTION SAMPLE] ...\n\n" + middle + "\n\n... [END SECTION] ...\n\n" + end


def process_files_intelligently(files_data: list) -> tuple:
    """
    Intelligently process multiple files with smart token distribution
    
    Args:
        files_data: List of dicts with 'name' and 'content' (bytes)
        
    Returns:
        Tuple of (processed_content_string, metadata_dict)
    """
    MAX_TOTAL_TOKENS = 120000  # Reserves room for: prompt template (~8k), images (~10k), output (16k), safety margin
    CHARS_PER_TOKEN = 4
    MAX_TOTAL_CHARS = MAX_TOTAL_TOKENS * CHARS_PER_TOKEN  # 480k chars total
    
    if not files_data:
        return "", {}
    
    # First pass: Extract and measure all files
    extracted_files = []
    for file_data in files_data:
        filename = file_data.get('name', 'unknown')
        file_content_bytes = file_data.get('content', b'')
        
        # Extract text
        extracted_text = process_file(filename, file_content_bytes, use_vision_api=False, max_chars_per_file=999999999)
        
        extracted_files.append({
            'name': filename,
            'content': extracted_text,
            'size': len(extracted_text),
            'tokens': len(extracted_text) // CHARS_PER_TOKEN
        })
    
    # Calculate total size
    total_size = sum(f['size'] for f in extracted_files)
    total_tokens = total_size // CHARS_PER_TOKEN
    
    logger.info(f"📊 File Analysis: {len(extracted_files)} files, {total_size:,} chars (~{total_tokens:,} tokens)")
    
    # If under limit, return everything
    if total_size <= MAX_TOTAL_CHARS:
        logger.info("✅ All files fit within token limit - no truncation needed")
        combined = "\n\n".join([f"### {f['name']}\n{f['content']}" for f in extracted_files])
        return combined, {
            'total_files': len(extracted_files),
            'total_chars': total_size,
            'total_tokens': total_tokens,
            'truncated': False
        }
    
    # Smart distribution: prioritize smaller files
    logger.info(f"⚠️ Total content ({total_tokens:,} tokens) exceeds limit ({MAX_TOTAL_TOKENS:,} tokens)")
    logger.info("🧠 Applying intelligent token distribution...")
    
    # Sort by size (smallest first - keep these intact)
    sorted_files = sorted(extracted_files, key=lambda x: x['size'])
    
    processed_files = []
    remaining_chars = MAX_TOTAL_CHARS
    
    for file_info in sorted_files:
        filename = file_info['name']
        content = file_info['content']
        size = file_info['size']
        
        if size <= remaining_chars:
            # File fits completely
            processed_files.append({
                'name': filename,
                'content': content,
                'truncated': False,
                'original_size': size,
                'final_size': size
            })
            remaining_chars -= size
            logger.info(f"  ✓ {filename}: {size:,} chars (complete)")
        else:
            # Need to truncate this file
            if remaining_chars > 10000:  # Only include if we have meaningful space
                truncated = smart_truncate(content, remaining_chars, filename)
                processed_files.append({
                    'name': filename,
                    'content': truncated,
                    'truncated': True,
                    'original_size': size,
                    'final_size': len(truncated)
                })
                logger.info(f"  ⚡ {filename}: {size:,} → {len(truncated):,} chars (truncated)")
                remaining_chars = 0
            else:
                # Not enough space, use placeholder
                placeholder = f"[Large file truncated due to token limits: {filename} - {size:,} chars]"
                processed_files.append({
                    'name': filename,
                    'content': placeholder,
                    'truncated': True,
                    'original_size': size,
                    'final_size': len(placeholder)
                })
                logger.info(f"  ⊘ {filename}: {size:,} chars (placeholder only)")
    
    # Combine all processed content
    combined = "\n\n".join([f"### {f['name']}\n{f['content']}" for f in processed_files])
    
    metadata = {
        'total_files': len(extracted_files),
        'files_complete': sum(1 for f in processed_files if not f['truncated']),
        'files_truncated': sum(1 for f in processed_files if f['truncated']),
        'original_total_chars': total_size,
        'original_total_tokens': total_tokens,
        'final_total_chars': len(combined),
        'final_total_tokens': len(combined) // CHARS_PER_TOKEN,
        'truncated': True
    }
    
    logger.info(f"📦 Final: {metadata['final_total_tokens']:,} tokens ({metadata['files_complete']} complete, {metadata['files_truncated']} truncated)")
    
    return combined, metadata


def process_files_with_vision(files_data: list) -> tuple:
    """
    Process files with Claude Vision API support for images
    
    Args:
        files_data: List of dicts with 'name' and 'content' (bytes or str)
        
    Returns:
        Tuple of (text_content_string, image_content_list, metadata_dict)
    """
    MAX_TOTAL_TOKENS = 120000  # Reserves room for: prompt template (~8k), images (~10k), output (16k), safety margin
    CHARS_PER_TOKEN = 4
    MAX_TOTAL_CHARS = MAX_TOTAL_TOKENS * CHARS_PER_TOKEN
    MAX_IMAGE_SIZE_MB = 5  # Claude API limit per image
    MAX_IMAGE_SIZE_BYTES = MAX_IMAGE_SIZE_MB * 1024 * 1024
    
    if not files_data:
        return "", [], {}
    
    # Separate images from text documents
    text_files = []
    image_files = []
    
    for file_data in files_data:
        filename = file_data.get('name', 'unknown')
        content = file_data.get('content', b'')
        file_extension = Path(filename).suffix.lower().lstrip('.')
        
        if file_extension in ['png', 'jpg', 'jpeg', 'gif', 'webp']:
            # This is an image
            # Check if content is already base64 string (from frontend) or bytes
            if isinstance(content, str):
                # Already base64 from frontend - strip data URI prefix if present
                base64_data = content
                # Remove data URI prefix like "data:image/png;base64,"
                if base64_data.startswith('data:'):
                    if ';base64,' in base64_data:
                        base64_data = base64_data.split(';base64,', 1)[1]
                        logger.info(f"🖼️  {filename}: Stripped data URI prefix from base64")
                    else:
                        logger.warning(f"⚠️  {filename}: Data URI found but no base64 marker")
                else:
                    logger.info(f"🖼️  {filename}: Using pre-encoded base64 from frontend")
                
                # Decode → resize → re-encode to reduce token consumption
                try:
                    raw_bytes = base64.b64decode(base64_data)
                    raw_bytes = resize_image_for_api(raw_bytes)
                    base64_data = base64.standard_b64encode(raw_bytes).decode('utf-8')
                    logger.info(f"🖼️  {filename}: Resized and re-encoded image from frontend")
                except Exception as resize_err:
                    logger.warning(f"⚠️  {filename}: Could not resize frontend image: {resize_err}, using original")
            else:
                # Binary bytes - resize then encode
                content = resize_image_for_api(content)
                base64_data = base64.standard_b64encode(content).decode('utf-8')
                logger.info(f"🖼️  {filename}: Resized and encoded binary to base64")
            
            # Validate base64 data
            if not base64_data or len(base64_data.strip()) == 0:
                logger.error(f"❌ {filename}: Empty base64 data, skipping")
                continue
            
            # Clean base64 data FIRST - remove any whitespace
            base64_data = base64_data.strip().replace('\n', '').replace('\r', '').replace(' ', '')
            
            # Validate that base64 can be decoded
            try:
                test_decode = base64.b64decode(base64_data, validate=True)
                actual_size = len(test_decode)
                logger.info(f"✓ {filename}: Valid base64, decoded to {actual_size / 1024:.1f}KB")
            except Exception as decode_err:
                logger.error(f"❌ {filename}: Invalid base64 data - {decode_err}")
                logger.error(f"   First 100 chars: '{base64_data[:100]}'")
                continue
            
            # Check actual decoded image size
            if actual_size > MAX_IMAGE_SIZE_BYTES:
                logger.warning(f"⚠️  {filename}: Image too large ({actual_size / 1024 / 1024:.1f}MB > {MAX_IMAGE_SIZE_MB}MB), skipping")
                continue
            
            # Validate it's actually an image by checking header
            try:
                from PIL import Image
                img = Image.open(io.BytesIO(test_decode))
                img.verify()
                logger.info(f"✓ {filename}: Valid {img.format} image, {img.size[0]}x{img.size[1]}px")
            except Exception as img_err:
                logger.error(f"❌ {filename}: Not a valid image file - {img_err}")
                continue
            
            # Determine media type
            media_type_map = {
                'jpg': 'image/jpeg',
                'jpeg': 'image/jpeg',
                'png': 'image/png',
                'gif': 'image/gif',
                'webp': 'image/webp'
            }
            media_type = media_type_map.get(file_extension.lower(), 'image/jpeg')
            
            # Log image details
            logger.info(f"📷 {filename}: {media_type}, {actual_size / 1024:.1f}KB, base64 length: {len(base64_data)}")
            
            image_files.append({
                'name': filename,
                'data': {
                    'type': 'image',
                    'source': {
                        'type': 'base64',
                        'media_type': media_type,
                        'data': base64_data
                    }
                },
                'type': 'image'
            })
        else:
            # Text document - extract content
            text_files.append(file_data)
    
    # Process text documents intelligently
    text_content = ""
    text_metadata = {}
    
    if text_files:
        text_content, text_metadata = process_files_intelligently(text_files)
    
    # Combine metadata
    metadata = {
        **text_metadata,
        'total_images': len(image_files),
        'image_files': [img['name'] for img in image_files]
    }
    
    logger.info(f"📊 Processing complete: {text_metadata.get('total_files', 0)} text docs, {len(image_files)} images")
    
    return text_content, image_files, metadata


def truncate_content(content: str, max_chars: int) -> str:
    """
    Truncate content to stay within character limits
    """
    if len(content) > max_chars:
        truncated = content[:max_chars]
        return f"{truncated}\n\n... [Content truncated - Original: {len(content):,} chars, Showing: {max_chars:,} chars]"
    return content
